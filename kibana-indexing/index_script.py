import openpyxl

###################
##### COLUMNS #####
###################

INDEX = 3
ORIGINAL_SIZE = 9
SCALED_SIZE = 11
HUB = 12

###################
##### SIZES #######
###################

INDEX_SIZE_GB = 'gb'
INDEX_SIZE_MB = 'mb'

##############################
###### HUBS & COUNTERS #######
##############################

hubs = {
    'EDB':['edb', 'containerlogs-edb'],
    'HOME':['home', 'containerlogs-home'],
    'OB':['ob', 'obcompete'],
    'MYNW':['mynw'],
    'RAAS':['raas'],
    'NWAP':['ndap', 'ops', 'containerlogs', 'telegraf', 'beat', 'tgw', 'watcher', 'prod', 'monitoring'],
}

counters = {
    'EDB': 0,
    'HOME': 0,
    'OB': 0,
    'MYNW': 0,
    'RAAS': 0,
    'NWAP': 0,
}

#/Users/Shahe.Islam/developer/ndap-journey/ndap-journey.xlsx
#/Users/Shahe.Islam/developer/ndap-journey/ndap-journey-test.xlsx

open_path = input("Input the file open path: ")

wb = openpyxl.load_workbook(open_path)
ws = wb['Sheet 1']

def list_headings():
    for k in hubs:
        wb.create_sheet(k)
        nws = wb[k]
        for i,row in enumerate(ws.iter_rows(max_row=1)):
            for j,col in enumerate(row):
                nws.cell(row=i+1,column=j+1).value = col.value

def index_size_scaler(size):
    if INDEX_SIZE_GB in size:
        size = size.strip(INDEX_SIZE_GB)
    elif INDEX_SIZE_MB in size:
        size = size.strip(INDEX_SIZE_MB)
        size = float(size)/1024
    else:
        size = 0
    return size

def sum_indexes():
    for k in hubs:
        ws = wb[k]
        data = [ws.cell(row=i,column=SCALED_SIZE).value for i in range(2, counters[k] + 2)]
        print(k + ": " + str(int(sum(data))) + ' GB')

for i in range(1, ws.max_row + 1):

    if i == 1:
        list_headings()
    else:
        size = ws.cell(row=i, column=ORIGINAL_SIZE).value
        scaled_size = index_size_scaler(size)
        ws.cell(row=i, column=SCALED_SIZE).value = float(scaled_size)

        index = ws.cell(row=i, column=INDEX).value

        for hub, journey in hubs.items():

            if any(x in index for x in journey):

                correctws = wb[hub]
                newrow = counters[hub] + 2

                for row in ws.iter_rows(min_row = i, max_row = i):
                    for j, cell in enumerate(row):
                        correctws.cell(row=newrow,column=j+1).value = cell.internal_value

                counters[hub] += 1
                break

            else:
                hub = 'default'

sum_indexes()

#/Users/Shahe.Islam/developer/ndap-journey/ndap-journey.xlsx
#/Users/Shahe.Islam/developer/ndap-journey/ndap-journey-test.xlsx

save_path = input("Input the file save path: ")
wb.save(save_path)
print("Your file has been saved at: " + save_path)
